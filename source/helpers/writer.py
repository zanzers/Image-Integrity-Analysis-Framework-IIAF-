import os
import json
from typing import Dict, Optional
import numpy as np
from pathlib import Path
import openpyxl
from openpyxl import Workbook, load_workbook


class FeatureJsonWriter:
    extracted_json_path = "ExtractedData"
    prediction_json_path = os.path.join("ExtractedData", "Output")

    @staticmethod
    def get_json_path(type_: str = "features") -> str:
        if type_ == "prediction":
            folder = FeatureJsonWriter.prediction_json_path
            file_name = "classification_result.json"
        else:
            folder = FeatureJsonWriter.extracted_json_path
            file_name = "features_result.json"

        if not os.path.exists(folder):
            os.makedirs(folder)
        return os.path.join(folder, file_name)

    @staticmethod
    def append_features(features: Dict[str, float], type_: str = "features"):
        if not features:
            print("[WARN] No features to save in JSON.")
            return
        try:
            json_path = FeatureJsonWriter.get_json_path(type_)
            folder = os.path.dirname(json_path)
            if os.path.exists(json_path) and os.path.isdir(json_path):
                print(f"[WARN] Found a directory with same name as file: {json_path}. Deleting it...")
                import shutil
                shutil.rmtree(json_path)
            if not os.path.exists(folder):
                os.makedirs(folder)
            if os.path.exists(json_path):
                os.remove(json_path)
            json_content = json.dumps(features, indent=4, default=FeatureJsonWriter._json_serializable)
            with open(json_path, 'w') as f:
                f.write(json_content)
            print(f"[INFO] Features successfully appended to JSON: {json_path}")
        except Exception as ex:
            print(f"[ERROR] Failed to write JSON file: {ex}")

    @staticmethod
    def _json_serializable(obj):
        if isinstance(obj, np.generic):
            return obj.item()
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        if isinstance(obj, (bytes, bytearray)):
            return obj.decode('utf-8')
        raise TypeError(f"Object of type {type(obj).__name__} is not JSON serializable")



class FeatureExcelWriter:
    base_folder = Path(__file__).parent.parent.parent / "ExtractedData"
    file_path = base_folder / "features_dataset.xlsx"
    is_first_write = True

    @staticmethod
    def append_features(features: Dict[str, float], label: Optional[str] = None):
        if not features:
            print("[WARN] No features to save in Excel.")
            return

        try:
            FeatureExcelWriter.base_folder.mkdir(parents=True, exist_ok=True)

            if FeatureExcelWriter.is_first_write or not FeatureExcelWriter.file_path.exists():
                if FeatureExcelWriter.file_path.exists():
                    FeatureExcelWriter.file_path.unlink()

                print(f"[INFO] Creating new Excel file at: {FeatureExcelWriter.file_path}")

                wb = Workbook()
                ws = wb.active
                ws.title = "Features"

              
                ws.append(["Label"] + list(features.keys()))

                
                ws.append([label or ""] + [f"{v:.6f}" for v in features.values()])

                wb.save(FeatureExcelWriter.file_path)

                FeatureExcelWriter.is_first_write = False
            else:
                wb = load_workbook(FeatureExcelWriter.file_path)
                ws = wb.active

                
                ws.append([label or ""] + [f"{v:.6f}" for v in features.values()])

                wb.save(FeatureExcelWriter.file_path)

            print(f"[INFO] Features successfully appended to Excel: {FeatureExcelWriter.file_path}")
        except Exception as ex:
            print(f"[ERROR] Failed to write Excel file: {ex}")


       
